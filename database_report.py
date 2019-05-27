# database_report.py

import openpyxl
import sys
book = openpyxl.load_workbook('Database.xlsx')

sheet = book.active

sheets = book.sheetnames
file = open("invoice.txt", "a")

print(f"All sheet names {book.sheetnames} ")

customers_list = []
invoices_list = []
lineItems_list = []
products_list = []
frequency = {}
frequency2 = {} # this is added so that the count start from 1 again on the product list because the list is created
values = []


class Build():
	@staticmethod
	def check_for_blanks():
		for sheet in book:
			for row in sheet.iter_rows():
				empty_cell_count = 0
				for cell in row:
					if cell.value == None:
						empty_cell_count += 1
						print(f"The {cell} is null.")
			print(f"End Check. {sheet}")
		print("----------------------------------------------------")

	@staticmethod
	def build_customer_list():
		# for row in book[sheets[0]].iter_rows():
		# 	for cell in row:
		# 		ref = book[sheets[0]].cell(row = cell.row, column = 2).value
		# 		if ref not in values:
		# 			values.append(ref)
		# print("Unique Values: ", values)
		for row in book[sheets[0]].iter_rows():
			for cell in row:
				obj = {}
				CustomerNumber = book[sheets[0]].cell(row = cell.row, column = 1).value
				CustomerName = book[sheets[0]].cell(row = cell.row, column = 2).value
				Address = book[sheets[0]].cell(row = cell.row, column = 3).value
				Phone = book[sheets[0]].cell(row = cell.row, column = 4).value
				Contact = book[sheets[0]].cell(row = cell.row, column = 5).value
				obj["CustomerNumber"] = CustomerNumber
				obj["CustomerName"] = CustomerName
				obj["Address"] = Address
				obj["Phone"] = Phone
				obj["Contact"] = Contact
			frequency[CustomerName] = frequency.get(CustomerName, 0) + 1
			frequency[CustomerNumber] = frequency.get(CustomerNumber, 0) + 1
			# print("Test the frequency of Customer ", frequency)
			if frequency[CustomerName] > 1 or frequency[CustomerNumber] > 1:
				print("Duplicate Customer Name: ", CustomerName)
				print("Duplicate Customer Number: ", CustomerNumber)
			else:
				customers_list.append(obj)
					
		# print("customers_list testing ", customers_list, "\n")
		# print("customers_list testing ", customers_list[5], "\n")
		return customers_list

	@staticmethod
	def build_invoices_list():
		for row in book[sheets[1]].iter_rows():
			for cell in row:
				obj = {}
				InvoiceNumber = book[sheets[1]].cell(row = cell.row, column = 1).value
				CustomerName = book[sheets[1]].cell(row = cell.row, column = 2).value
				Date = book[sheets[1]].cell(row = cell.row, column = 3).value
				obj["InvoiceNumber"] = InvoiceNumber
				obj["CustomerNumber"] = CustomerName
				obj["Date"] = Date
			frequency[InvoiceNumber] = frequency.get(InvoiceNumber, 0) + 1
			if frequency[InvoiceNumber] > 1:
				print("Duplicate Invoice Number: ", InvoiceNumber)
			else:
				invoices_list.append(obj)

		# print("invoices_list testing ", invoices_list, "\n")	
		# print("invoices_list testing ", invoices_list[5], "\n")
		return invoices_list

	@staticmethod
	def build_lineItems_list():
		for row in book[sheets[2]].iter_rows():
			for cell in row:
				obj = {}
				obj["InvoiceNumber"] = book[sheets[2]].cell(row = cell.row, column = 1).value
				obj["ProductNumber"] = book[sheets[2]].cell(row = cell.row, column = 2).value
				obj["Units"] = book[sheets[2]].cell(row = cell.row, column = 3).value

			lineItems_list.append(obj)
		# print("lineItems_list testing ", lineItems_list, "\n")
		# print("lineItems_list testing ", lineItems_list[5], "\n")
		return lineItems_list

	@staticmethod
	def build_products_list():
		for row in book[sheets[3]].iter_rows():
			for cell in row:
				obj = {}
				ProductNumber = book[sheets[3]].cell(row = cell.row, column = 1).value
				ProductName = book[sheets[3]].cell(row = cell.row, column = 2).value
				UnitCost = book[sheets[3]].cell(row = cell.row, column = 3).value
				obj["ProductNumber"] = ProductNumber
				obj["ProductName"] = ProductName
				obj["UnitCost"] = UnitCost
			frequency2[ProductNumber] = frequency2.get(ProductNumber, 0) + 1
			frequency2[ProductName] = frequency2.get(ProductName, 0) + 1
			# print("test the Product: ", ProductNumber, ProductName, frequency[ProductNumber])
			if frequency2[ProductName] > 1 or frequency2[ProductNumber] > 1:
				print("Duplicate Product Name: ", ProductName)
				print("Duplicate Product Number: ", ProductNumber)
			else:
				products_list.append(obj)
			
		# print("products_list testing ", products_list, "\n")
		# print("products_list testing ", products_list[5], "\n")
		return products_list

	def check_for_numeric():
		for sheet in book:
			for row in sheet.iter_rows():
				# numeric_count = 0
				for cell in row:
					if str(cell.value).isnumeric() == True:
						# numeric_count += 1
						print("The cell is number: ", cell.value)
					else:
						print("The cell is not a number: ", cell.value)
				# print("the numeric count is: ", numeric_count)

	# @staticmethod
	def build_the_master_list():
		Build.build_customer_list()
		Build.build_invoices_list()
		Build.build_lineItems_list()
		Build.build_products_list()
		
Build()
# Build.check_for_numeric()
Build.check_for_blanks()
Build.build_the_master_list()


sys.stdout = open("invoice.txt", "wt")
class Create_invoice():
	def create_invoice(inv):
		for obj in invoices_list:
			if obj["InvoiceNumber"] == int(inv):
				customer_referance = obj["CustomerNumber"]
				print("Invoice Number: 	", inv)
				print("------------")
				print("Customer ID:    	", customer_referance)
				print("Invoice Number: 	", inv)
				print("Date:           	", obj["Date"])
				print("------------------------------------")

		for obj in customers_list:
			if obj["CustomerNumber"] == customer_referance:
				# print("------------")
				print("Customer Name: 		", obj["CustomerName"])
				print("Address: 			", obj["Address"])
				print("Phone: 				", obj["Phone"])
				print("Contact: 			", obj["Contact"])
				print("------------------------------------")

		grand_total = 0
		for obj in lineItems_list:
			if obj["InvoiceNumber"] == int(inv):
				product_referance = obj["ProductNumber"]
				product_purchase = obj["Units"]
				for obj in products_list:
					if obj["ProductNumber"] == product_referance:
						total = "{:.2f}".format(product_purchase * obj["UnitCost"])
						grand_total += float(total)
						# print("------------")
						print("Product Name: ", obj["ProductName"], " Number: ", product_purchase, " Cost: ", "{:.2f}".format(obj["UnitCost"]), " Total: ", total)
						print("------------------------------------------------------------------------")
		print("											Grand total: ", grand_total)
		# return "Done"


file.close()
Create_invoice.create_invoice(12)
print("\n")
Create_invoice.create_invoice(16)
print("\n")
Create_invoice.create_invoice(8)